"""Report export service for PDF, Excel, and CSV formats."""

import io
import json
from datetime import datetime
from typing import Any, Dict, List, Optional
from uuid import UUID
import csv


class ReportExporter:
    """Service for exporting reports in various formats."""

    @staticmethod
    def export_to_csv(report_data: Dict[str, Any], filename: str) -> bytes:
        """Export report data to CSV format."""
        output = io.StringIO()
        
        # Write report header
        output.write(f"Report: {report_data.get('report_type', 'Unknown')}\n")
        if report_data.get('period'):
            output.write(f"Period: {report_data['period'].get('start')} to {report_data['period'].get('end')}\n")
        output.write(f"Generated: {datetime.utcnow().isoformat()}\n\n")

        # Write data
        data = report_data.get('data', [])
        if not data:
            output.write("No data available\n")
            return output.getvalue().encode('utf-8')

        writer = csv.DictWriter(output, fieldnames=data[0].keys() if data else [])
        writer.writeheader()
        writer.writerows(data)

        # Write summary
        output.write("\n\nSummary:\n")
        summary = report_data.get('summary', {})
        for key, value in summary.items():
            output.write(f"{key}: {value}\n")

        return output.getvalue().encode('utf-8')

    @staticmethod
    def export_to_json(report_data: Dict[str, Any], filename: str) -> bytes:
        """Export report data to JSON format."""
        json_data = json.dumps(report_data, indent=2, default=str)
        return json_data.encode('utf-8')

    @staticmethod
    def export_to_excel(report_data: Dict[str, Any], filename: str) -> bytes:
        """Export report data to Excel format."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export")

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Write header
        row = 1
        ws[f"A{row}"] = f"Report: {report_data.get('report_type', 'Unknown')}"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        
        row += 1
        if report_data.get('period'):
            ws[f"A{row}"] = f"Period: {report_data['period'].get('start')} to {report_data['period'].get('end')}"
        
        row += 1
        ws[f"A{row}"] = f"Generated: {datetime.utcnow().isoformat()}"
        
        row += 2

        # Write data
        data = report_data.get('data', [])
        if data:
            # Write headers
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            row += 1

            # Write data rows
            for data_row in data:
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col_idx)
                    value = data_row.get(header)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                row += 1

        # Write summary
        row += 2
        ws[f"A{row}"] = "Summary"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        
        row += 1
        summary = report_data.get('summary', {})
        for key, value in summary.items():
            ws[f"A{row}"] = str(key)
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Write to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_to_pdf(report_data: Dict[str, Any], filename: str) -> bytes:
        """Export report data to PDF format."""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.units import inch
            from io import BytesIO
        except ImportError:
            raise ImportError("reportlab is required for PDF export")

        # Create PDF document
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor("#1F2937"),
            spaceAfter=12,
        )
        elements.append(Paragraph(f"Report: {report_data.get('report_type', 'Unknown')}", title_style))
        
        # Period info
        if report_data.get('period'):
            period_text = f"Period: {report_data['period'].get('start')} to {report_data['period'].get('end')}"
            elements.append(Paragraph(period_text, styles['Normal']))
        
        elements.append(Paragraph(f"Generated: {datetime.utcnow().isoformat()}", styles['Normal']))
        elements.append(Spacer(1, 0.3 * inch))

        # Data table
        data = report_data.get('data', [])
        if data:
            table_data = []
            # Headers
            headers = list(data[0].keys())
            table_data.append(headers)
            
            # Data rows
            for row in data:
                table_data.append([str(row.get(h, '')) for h in headers])

            table = Table(table_data, colWidths=[1.2 * inch] * len(headers))
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F2937")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            elements.append(table)

        # Summary
        elements.append(Spacer(1, 0.2 * inch))
        summary_title = ParagraphStyle(
            'SummaryTitle',
            parent=styles['Heading2'],
            fontSize=12,
        )
        elements.append(Paragraph("Summary", summary_title))
        
        summary = report_data.get('summary', {})
        for key, value in summary.items():
            elements.append(Paragraph(f"<b>{key}:</b> {value}", styles['Normal']))

        # Build PDF
        doc.build(elements)
        return output.getvalue()


class ExportService:
    """Main export service orchestrator."""

    EXPORTERS = {
        'csv': ReportExporter.export_to_csv,
        'json': ReportExporter.export_to_json,
        'excel': ReportExporter.export_to_excel,
        'pdf': ReportExporter.export_to_pdf,
    }

    @classmethod
    def export_report(
        cls,
        report_data: Dict[str, Any],
        export_format: str,
        filename: str,
    ) -> tuple[bytes, str]:
        """Export report data in specified format.
        
        Returns:
            Tuple of (file_bytes, mime_type)
        """
        if export_format not in cls.EXPORTERS:
            raise ValueError(f"Unsupported export format: {export_format}")

        exporter = cls.EXPORTERS[export_format]
        file_bytes = exporter(report_data, filename)

        mime_types = {
            'csv': 'text/csv',
            'json': 'application/json',
            'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'pdf': 'application/pdf',
        }

        return file_bytes, mime_types.get(export_format, 'application/octet-stream')

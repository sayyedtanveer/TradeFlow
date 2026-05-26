from __future__ import annotations

import csv
import io
import uuid
from datetime import datetime, timezone
from difflib import SequenceMatcher
from typing import Any, Dict, List, Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select

from backend.app.infrastructure.persistence.models.material_model import MaterialModel
from backend.app.infrastructure.persistence.models.unit_of_measure_model import UnitOfMeasureModel
from backend.app.infrastructure.persistence.models.material_category_model import MaterialCategoryModel
from backend.app.interfaces.api.v1.dependencies.auth import (
    get_container,
    get_current_tenant_id,
    get_current_user_id,
)
from backend.app.interfaces.api.v1.dependencies.permissions import require_permission

router = APIRouter(prefix="/inventory/material-onboarding", tags=["Material Onboarding"])

# ── In-memory session store ───────────────────────────────────────────────────
_sessions: Dict[str, Dict[str, Any]] = {}

# ── Target field definitions ──────────────────────────────────────────────────
RAW_MATERIAL_COLUMNS = [
    "item_code", "material_name", "material_category", "material_type", "uom",
    "batch_tracking_enabled", "shelf_life", "expiry_tracking", "warehouse", "zone", "rack_bin",
    "min_stock", "max_stock", "reorder_level", "reorder_quantity", "barcode", "traceability_enabled",
    "qc_required", "approved_supplier", "supplier_item_code", "purchase_uom", "lead_time", "moq",
    "length_uom", "cuttable_inventory", "remaining_quantity_tracking", "decimal_precision", "reusable_remainder",
]

# Fields that are considered protected (changing them on an existing material requires confirmation)
PROTECTED_FIELDS = {"material_type", "batch_tracking_enabled", "traceability_enabled", "uom"}

# Map target field → MaterialModel attribute
FIELD_TO_MODEL: Dict[str, str] = {
    "item_code": "item_code",
    "material_name": "name",
    "material_type": "material_type",
    "batch_tracking_enabled": "is_batch_tracked",
    "min_stock": "safety_stock",
    "reorder_level": "reorder_level",
    "lead_time": "lead_time_days",
    "length_uom": "length_uom",
    "qc_required": "qc_required_flag",
    "traceability_enabled": "is_serialized",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower().replace(" ", "_"), b.lower().replace(" ", "_")).ratio()


def _suggest_mapping(headers: List[str]) -> Dict[str, str]:
    """Auto-suggest column-to-field mappings based on name similarity."""
    mapping: Dict[str, str] = {}
    for header in headers:
        best_field, best_score = "", 0.0
        for field in RAW_MATERIAL_COLUMNS:
            score = _similarity(header, field)
            if score > best_score:
                best_score, best_field = score, field
        if best_score >= 0.6:
            mapping[header] = best_field
    return mapping


def _parse_file(file_bytes: bytes, file_name: str) -> tuple[List[str], List[Dict[str, str]]]:
    """Return (headers, rows) from an xlsx or csv file."""
    name_lower = (file_name or "").lower()
    if name_lower.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        headers = list(reader.fieldnames or [])
        rows = [{k: (str(v).strip() if v else "") for k, v in row.items()} for row in reader]
        return headers, rows
    else:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                raise ValueError("No active worksheet found in workbook")
            headers: List[str] = []
            rows: List[Dict[str, str]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).strip() if c is not None else f"col_{j}" for j, c in enumerate(row)]
                    continue
                if all(c is None for c in row):
                    continue
                rows.append({headers[j]: (str(c).strip() if c is not None else "") for j, c in enumerate(row) if j < len(headers)})
            wb.close()
            return headers, rows
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"Could not parse file: {exc}")


def _apply_mapping(raw_row: Dict[str, str], mapping: Dict[str, str]) -> Dict[str, Any]:
    """Convert a raw row using the column mapping."""
    data: Dict[str, Any] = {}
    for source, target in mapping.items():
        if target and target != "_ignore":
            data[target] = raw_row.get(source, "")
    return data


def _bool_val(v: Any) -> bool:
    return str(v).strip().lower() in ("true", "yes", "1", "y")


def _validate_row(
    data: Dict[str, Any],
    row_num: int,
    existing_codes: Dict[str, Dict[str, Any]],
) -> tuple[str, str, List[Dict], List[Dict]]:
    """
    Returns (classification, status, issues, protected_changes).
    classification: 'new' | 'update' | 'skip'
    """
    issues: List[Dict] = []
    protected_changes: List[Dict] = []

    name = str(data.get("material_name", "")).strip()
    item_code = str(data.get("item_code", "")).strip()
    uom = str(data.get("uom", "")).strip()
    category = str(data.get("material_category", "")).strip()

    # Required field checks
    if not name:
        issues.append({"field": "material_name", "severity": "error", "message": "Material name is required"})
    if not uom:
        issues.append({"field": "uom", "severity": "error", "message": "Unit of measure (uom) is required"})
    if not category:
        issues.append({"field": "material_category", "severity": "warning", "message": "No category specified — will be left blank"})

    # Numeric validation
    for field in ("reorder_level", "min_stock", "max_stock", "lead_time"):
        val = data.get(field, "")
        if val:
            try:
                float(str(val).replace(",", ""))
            except ValueError:
                issues.append({"field": field, "severity": "error", "message": f"{field} must be a number"})

    # Determine classification and protected changes
    lookup_key = item_code.upper() if item_code else name.upper()
    existing = existing_codes.get(lookup_key)

    if existing:
        classification = "update"
        # Check protected fields
        for pf in PROTECTED_FIELDS:
            new_val = str(data.get(pf, "")).strip()
            if not new_val:
                continue
            model_attr = FIELD_TO_MODEL.get(pf, pf)
            old_val = str(existing.get(model_attr, "")).strip()
            if old_val and old_val != new_val:
                protected_changes.append({"field": pf, "from": old_val, "to": new_val})
    else:
        classification = "new"

    has_errors = any(i["severity"] == "error" for i in issues)
    row_status = "error" if has_errors else "ready"
    if has_errors:
        classification = "skip"

    return classification, row_status, issues, protected_changes


# ── Schemas ───────────────────────────────────────────────────────────────────

class ValidateRequest(BaseModel):
    mapping: Dict[str, str]


class ExecuteRequest(BaseModel):
    dry_run: bool = False


class RowUpdateRequest(BaseModel):
    class Config:
        extra = "allow"


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get(
    "/template",
    summary="Download the material onboarding template",
    # No auth required — public template
)
async def get_template(format: str = Query(default="csv", pattern="^(csv|xlsx)$")):
    """Return a downloadable template file with the expected column headers."""
    if format == "xlsx":
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                raise ValueError("Could not create active worksheet")
            ws.title = "Materials"
            ws.append(RAW_MATERIAL_COLUMNS)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=material-onboarding-template.xlsx"},
            )
        except ImportError:
            pass  # fall through to csv

    # CSV
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(RAW_MATERIAL_COLUMNS)
    writer.writerow([
        "MAT-001", "Steel Rod 10mm", "Raw Materials", "raw", "KG",
        "false", "", "", "", "", "",
        "10", "500", "20", "50", "", "false",
        "false", "", "", "", "7", "",
        "", "", "", "2", "",
    ])
    buf.seek(0)
    return StreamingResponse(
        io.BytesIO(buf.read().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=material-onboarding-template.csv"},
    )


@router.post(
    "/sessions",
    status_code=status.HTTP_201_CREATED,
    summary="Upload a file and create an onboarding session",
    dependencies=[Depends(require_permission("inventory:write"))],
)
async def create_session(
    file: UploadFile = File(...),
    tenant_id: uuid.UUID = Depends(get_current_tenant_id),
    user_id: uuid.UUID = Depends(get_current_user_id),
):
    """
    Parse the uploaded file and return session_id, column headers, and
    a JSON-stringified auto-suggested column mapping.
    """
    import json

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    headers, raw_rows = _parse_file(file_bytes, file.filename or "upload")

    if not headers:
        raise HTTPException(status_code=422, detail="No header row found in file")
    if not raw_rows:
        raise HTTPException(status_code=422, detail="No data rows found. Ensure row 1 is a header and data starts from row 2.")

    suggested = _suggest_mapping(headers)

    session_id = str(uuid.uuid4())
    _sessions[session_id] = {
        "session_id": session_id,
        "tenant_id": str(tenant_id),
        "user_id": str(user_id),
        "file_name": file.filename or "upload",
        "headers": headers,
        "raw_rows": raw_rows,
        "mapping": {},
        "rows": [],
        "protected_confirmed": False,
        "status": "uploaded",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    return {
        "session_id": session_id,
        "headers": headers,
        "mapping": json.dumps(suggested),
    }


@router.post(
    "/sessions/{session_id}/validate",
    summary="Apply column mapping and validate all rows",
    dependencies=[Depends(require_permission("inventory:write"))],
)
async def validate_session(
    session_id: str,
    body: ValidateRequest,
    request: Request,
    tenant_id: uuid.UUID = Depends(get_current_tenant_id),
):
    """Apply column mapping, run validation rules, and classify each row."""
    session = _get_session(session_id, tenant_id)

    container = get_container(request)
    existing_codes: Dict[str, Dict[str, Any]] = {}
    try:
        async with container.db_session_factory() as db:
            result = await db.execute(
                select(
                    MaterialModel.code,
                    MaterialModel.item_code,
                    MaterialModel.name,
                    MaterialModel.material_type,
                    MaterialModel.is_batch_tracked,
                    MaterialModel.is_serialized,
                    MaterialModel.safety_stock,
                    MaterialModel.reorder_level,
                ).where(
                    MaterialModel.tenant_id == tenant_id,
                    MaterialModel.is_deleted.is_(False),
                )
            )
            for row in result.fetchall():
                rec = dict(row._mapping)
                key = (row.item_code or row.code or "").upper()
                if key:
                    existing_codes[key] = rec
                name_key = (row.name or "").upper()
                if name_key:
                    existing_codes[name_key] = rec
    except Exception:
        pass

    mapping = body.mapping
    processed_rows = []
    for i, raw_row in enumerate(session["raw_rows"], start=2):
        row_id = str(uuid.uuid4())
        data = _apply_mapping(raw_row, mapping)
        classification, row_status, issues, protected_changes = _validate_row(data, i, existing_codes)
        processed_rows.append({
            "id": row_id,
            "row_number": i,
            "classification": classification,
            "status": row_status,
            "data": data,
            "issues": issues,
            "protected_changes": protected_changes,
        })

    session["mapping"] = mapping
    session["rows"] = processed_rows
    session["status"] = "validated"

    return {"validated": len(processed_rows)}


@router.get(
    "/sessions/{session_id}/preview",
    summary="Get validation preview for the session",
    dependencies=[Depends(require_permission("inventory:write"))],
)
async def get_preview(
    session_id: str,
    tenant_id: uuid.UUID = Depends(get_current_tenant_id),
):
    """Return per-row validation results and a summary."""
    session = _get_session(session_id, tenant_id)
    rows = session.get("rows", [])

    if not rows:
        raise HTTPException(status_code=400, detail="Session has not been validated yet. Call /validate first.")

    total = len(rows)
    new_count = sum(1 for r in rows if r["classification"] == "new")
    update_count = sum(1 for r in rows if r["classification"] == "update")
    skip_count = sum(1 for r in rows if r["classification"] == "skip")
    error_count = sum(1 for r in rows if r["status"] == "error")
    warning_count = sum(1 for r in rows for i in r["issues"] if i["severity"] == "warning")

    return {
        "summary": {
            "total_rows": total,
            "new": new_count,
            "update": update_count,
            "skip": skip_count,
            "errors": error_count,
            "warnings": warning_count,
        },
        "rows": rows,
    }


@router.post(
    "/sessions/{session_id}/confirm-protected",
    summary="Confirm that protected field changes are intentional",
    dependencies=[Depends(require_permission("inventory:write"))],
)
async def confirm_protected(
    session_id: str,
    tenant_id: uuid.UUID = Depends(get_current_tenant_id),
):
    session = _get_session(session_id, tenant_id)
    session["protected_confirmed"] = True
    return {"confirmed": True}


@router.post(
    "/sessions/{session_id}/execute",
    summary="Execute the import (create/update materials)",
    dependencies=[Depends(require_permission("inventory:write"))],
)
async def execute_session(
    session_id: str,
    body: ExecuteRequest,
    request: Request,
    tenant_id: uuid.UUID = Depends(get_current_tenant_id),
    user_id: uuid.UUID = Depends(get_current_user_id),
):
    """
    Commit the onboarding session. With dry_run=True runs all logic without
    writing to the database.
    """
    session = _get_session(session_id, tenant_id)
    rows = session.get("rows", [])
    if not rows:
        raise HTTPException(status_code=400, detail="Session has not been validated yet.")

    actionable = [r for r in rows if r["classification"] in ("new", "update") and r["status"] != "error"]

    created = updated = skipped = errors = 0
    error_messages: List[str] = []

    if not body.dry_run:
        container = get_container(request)

        # Fetch UOM and category lookups
        uom_map: Dict[str, uuid.UUID] = {}
        cat_map: Dict[str, uuid.UUID] = {}
        existing_map: Dict[str, MaterialModel] = {}

        try:
            async with container.db_session_factory() as db:
                uom_result = await db.execute(
                    select(UnitOfMeasureModel.code, UnitOfMeasureModel.id).where(
                        UnitOfMeasureModel.tenant_id == tenant_id,
                        UnitOfMeasureModel.is_deleted.is_(False),
                    )
                )
                for r in uom_result:
                    uom_map[r.code.upper()] = r.id

                cat_result = await db.execute(
                    select(MaterialCategoryModel.name, MaterialCategoryModel.id).where(
                        MaterialCategoryModel.tenant_id == tenant_id,
                        MaterialCategoryModel.is_deleted.is_(False),
                    )
                )
                for r in cat_result:
                    cat_map[r.name.upper()] = r.id

                mat_result = await db.execute(
                    select(MaterialModel).where(
                        MaterialModel.tenant_id == tenant_id,
                        MaterialModel.is_deleted.is_(False),
                    )
                )
                for mat in mat_result.scalars():
                    key = (mat.item_code or mat.code or "").upper()
                    if key:
                        existing_map[key] = mat
                    name_key = (mat.name or "").upper()
                    if name_key:
                        existing_map[name_key] = mat

                # Auto-create missing UOMs following project standard pattern
                for row in actionable:
                    data = row["data"]
                    uom_code = str(data.get("uom", "")).strip().upper()
                    if uom_code and uom_code not in uom_map:
                        new_uom = UnitOfMeasureModel(
                            id=uuid.uuid4(),
                            tenant_id=tenant_id,
                            code=uom_code,
                            name=uom_code,
                            precision=2,
                            is_active=True,
                            is_deleted=False,
                            created_at=datetime.now(timezone.utc),
                            updated_at=datetime.now(timezone.utc),
                        )
                        db.add(new_uom)
                        uom_map[uom_code] = new_uom.id

                # Auto-create missing categories following project standard pattern
                for row in actionable:
                    data = row["data"]
                    cat_name = str(data.get("material_category", "")).strip().upper()
                    if cat_name and cat_name not in cat_map:
                        new_cat = MaterialCategoryModel(
                            id=uuid.uuid4(),
                            tenant_id=tenant_id,
                            name=cat_name,
                            code_prefix=cat_name[:3] if len(cat_name) >= 3 else cat_name,
                            description=None,
                            is_active=True,
                            is_deleted=False,
                            created_at=datetime.now(timezone.utc),
                            updated_at=datetime.now(timezone.utc),
                        )
                        db.add(new_cat)
                        cat_map[cat_name] = new_cat.id
                
                await db.flush()
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Failed to auto-create master data: {exc}")

        async with container.db_session_factory() as db:
            for row in actionable:
                try:
                    data = row["data"]
                    name = str(data.get("material_name", "")).strip()
                    item_code = str(data.get("item_code", "")).strip()
                    uom_code = str(data.get("uom", "")).strip().upper()
                    cat_name = str(data.get("material_category", "")).strip().upper()

                    uom_id = uom_map.get(uom_code)
                    cat_id = cat_map.get(cat_name)

                    def _f(key: str) -> str | None:
                        val = data.get(key, "").strip()
                        return val if val else None

                    def _bool_val(val: str) -> bool:
                        return str(val).lower() in ("true", "1", "yes")

                    def _float_f(key: str) -> float | None:
                        try:
                            val = float(data.get(key, 0) or 0)
                            return val if val > 0 else None
                        except (TypeError, ValueError):
                            return None

                    def _int_f(key: str) -> int | None:
                        try:
                            val = int(data.get(key, 0) or 0)
                            return val if val > 0 else None
                        except (TypeError, ValueError):
                            return None

                    existing = existing_map.get(item_code.upper()) or existing_map.get(name.upper())
                    if not existing:
                        # Create
                        mat = MaterialModel(
                            id=uuid.uuid4(),
                            tenant_id=tenant_id,
                            code=item_code,
                            name=name,
                            base_unit_id=uom_id,
                            category_id=cat_id,
                            is_batch_tracked=_bool_val(data.get("batch_tracking_enabled", "false")),
                            is_serialized=_bool_val(data.get("traceability_enabled", "false")),
                            qc_required_flag=_bool_val(data.get("qc_required", "false")),
                            reorder_level=_float_f("reorder_level"),
                            safety_stock=_float_f("min_stock"),
                            lead_time_days=_int_f("lead_time"),
                            length_uom=_f("length_uom"),
                            current_cost=0,
                            current_stock=0,
                            reserved_stock=0,
                            is_active=True,
                            is_deleted=False,
                            created_by=user_id,
                            updated_by=user_id,
                        )
                        db.add(mat)
                        created += 1
                    else:
                        # Update
                        if session["protected_confirmed"]:
                            material_type = _f("material_type")
                            if material_type:
                                existing.material_type = material_type
                            existing.is_batch_tracked = _bool_val(data.get("batch_tracking_enabled", str(existing.is_batch_tracked)))
                            existing.is_serialized = _bool_val(data.get("traceability_enabled", str(existing.is_serialized)))
                            base_unit = uom_id or existing.base_unit_id
                            if base_unit:
                                existing.base_unit_id = base_unit
                        category = cat_id or existing.category_id
                        if category:
                            existing.category_id = category
                        reorder = _float_f("reorder_level")
                        if reorder:
                            existing.reorder_level = reorder
                        safety = _float_f("min_stock")
                        if safety:
                            existing.safety_stock = safety
                        lead = _int_f("lead_time")
                        if lead:
                            existing.lead_time_days = lead
                        length = _f("length_uom")
                        if length:
                            existing.length_uom = length
                        existing.updated_by = user_id
                        updated += 1
                except Exception as exc:
                    errors += 1
                    error_messages.append(f"Row {row['row_number']}: {exc}")

            skipped = len(rows) - len(actionable)
            try:
                await db.commit()
            except Exception as exc:
                await db.rollback()
                raise HTTPException(status_code=500, detail=f"Commit failed: {exc}")

        _sessions.pop(session_id, None)
    else:
        # Dry run: just count
        created = sum(1 for r in actionable if r["classification"] == "new")
        updated = sum(1 for r in actionable if r["classification"] == "update")
        skipped = len(rows) - len(actionable)

    return {
        "dry_run": body.dry_run,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "error_messages": error_messages,
    }


@router.patch(
    "/rows/{row_id}",
    summary="Correct a specific row's field values",
    dependencies=[Depends(require_permission("inventory:write"))],
)
async def update_row(
    row_id: str,
    request: Request,
    tenant_id: uuid.UUID = Depends(get_current_tenant_id),
):
    """Update a specific row's data fields and re-validate it."""
    body = await request.json()

    # Find the row across all sessions belonging to this tenant
    target_session = None
    target_row = None
    for sess in _sessions.values():
        if sess["tenant_id"] != str(tenant_id):
            continue
        for row in sess.get("rows", []):
            if row["id"] == row_id:
                target_session = sess
                target_row = row
                break
        if target_row:
            break

    if not target_row:
        raise HTTPException(status_code=404, detail="Row not found or session expired")

    # Apply updates
    for field, value in body.items():
        target_row["data"][field] = value

    # Re-validate the row
    existing_codes: Dict[str, Dict[str, Any]] = {}
    try:
        container = get_container(request)
        async with container.db_session_factory() as db:
            result = await db.execute(
                select(
                    MaterialModel.code, MaterialModel.item_code, MaterialModel.name,
                    MaterialModel.material_type, MaterialModel.is_batch_tracked,
                    MaterialModel.is_serialized, MaterialModel.safety_stock, MaterialModel.reorder_level,
                ).where(
                    MaterialModel.tenant_id == tenant_id,
                    MaterialModel.is_deleted.is_(False),
                )
            )
            for row in result.fetchall():
                rec = dict(row._mapping)
                key = (row.item_code or row.code or "").upper()
                if key:
                    existing_codes[key] = rec
    except Exception:
        pass

    classification, row_status, issues, protected_changes = _validate_row(
        target_row["data"], target_row["row_number"], existing_codes
    )
    target_row["classification"] = classification
    target_row["status"] = row_status
    target_row["issues"] = issues
    target_row["protected_changes"] = protected_changes

    return target_row


@router.get(
    "/sessions/{session_id}/validation-report",
    summary="Download a CSV validation report for the session",
    dependencies=[Depends(require_permission("inventory:write"))],
)
async def validation_report(
    session_id: str,
    tenant_id: uuid.UUID = Depends(get_current_tenant_id),
):
    session = _get_session(session_id, tenant_id)
    rows = session.get("rows", [])

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["row_number", "classification", "status", "material_name", "item_code", "uom", "issues"])
    for row in rows:
        data = row.get("data", {})
        issues_text = "; ".join(f"[{i['severity']}] {i['field']}: {i['message']}" for i in row.get("issues", []))
        writer.writerow([
            row["row_number"],
            row["classification"],
            row["status"],
            data.get("material_name", ""),
            data.get("item_code", ""),
            data.get("uom", ""),
            issues_text,
        ])

    buf.seek(0)
    return StreamingResponse(
        io.BytesIO(buf.read().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=material-onboarding-validation-{session_id}.csv"},
    )


# ── Internal helpers ──────────────────────────────────────────────────────────

def _get_session(session_id: str, tenant_id: uuid.UUID) -> Dict[str, Any]:
    session = _sessions.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Onboarding session not found or expired")
    if session["tenant_id"] != str(tenant_id):
        raise HTTPException(status_code=403, detail="Access denied")
    return session
